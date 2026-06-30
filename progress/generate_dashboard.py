"""
WizAccountant Project Dashboard Generator
Reads progress/pm-data.json and writes progress/WizAccountant-Dashboard.xlsx
Run after every PM check that processes a comment.
"""
import json, os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint, SeriesLabel

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE  = os.path.join(SCRIPT_DIR, "pm-data.json")
OUT_FILE   = os.path.join(SCRIPT_DIR, "WizAccountant-Dashboard.xlsx")

with open(DATA_FILE, encoding="utf-8") as f:
    D = json.load(f)

S   = D["summary"]
tasks    = D["tasks"]
burndown = D["burndown"]
billing  = D["billing_log"]

# ── Colours ──────────────────────────────────────────────────────────────────
C_DARK   = "1F3864"
C_MID    = "2E75B6"
C_LIGHT  = "BDD7EE"
C_GOLD   = "C6A84B"
C_GREEN  = "375623"
C_RED    = "C00000"
C_AMBER  = "ED7D31"
C_WHITE  = "FFFFFF"
C_LGREY  = "F2F2F2"
C_DGREY  = "595959"

STATUS_COLORS = {
    "Completed":     "375623",
    "In Progress":   "2E75B6",
    "Blocked":       "C00000",
    "Pending QA":    "7030A0",
    "Not Started":   "808080",
}
DELAY_COLORS = {
    "On Track":  "E2EFDA",
    "At Risk":   "FFF2CC",
    "Overdue":   "FCE4D6",
}
BILLING_COLORS = {
    "Within Estimate":   "E2EFDA",
    "Pending PJ Approval": "FFF2CC",
    "Approved":          "E2EFDA",
    "Rejected":          "FCE4D6",
    "Not Started":       "F2F2F2",
}

thin  = Side(style="thin",   color="CCCCCC")
med   = Side(style="medium", color="9E9E9E")
TB = Border(left=thin, right=thin, top=thin, bottom=thin)
MB = Border(left=med,  right=med,  top=med,  bottom=med)

def fill(c): return PatternFill("solid", fgColor=c)
def ctr():   return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def rgt():   return Alignment(horizontal="right",  vertical="center")

wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Dashboard"
ws.sheet_properties.tabColor = C_MID
ws.sheet_view.showGridLines = False

# Title
ws.merge_cells("A1:N1")
ws["A1"] = "WizAccountant / WizGate — Project Dashboard"
ws["A1"].font      = Font("Arial", size=16, bold=True, color=C_WHITE)
ws["A1"].fill      = fill(C_DARK)
ws["A1"].alignment = ctr()
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:N2")
ws["A2"] = f"Last updated: {D['last_updated']}   |   Developer: Varun   |   PM: Claude"
ws["A2"].font      = Font("Arial", size=10, italic=True, color=C_WHITE)
ws["A2"].fill      = fill(C_MID)
ws["A2"].alignment = ctr()
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 8

# ── KPI cards ────────────────────────────────────────────────────────────────
kpis = [
    ("Total Tasks",    S["total_tasks"],                    C_DARK),
    ("Completed",      S["completed"],                      "375623"),
    ("In Progress",    S["in_progress"],                    C_MID),
    ("Blocked",        S["blocked"],                        C_RED),
    ("Est. Hours",     f"{S['total_estimated_hours']} hrs", C_DGREY),
    ("Hours Used",     f"{S['hours_consumed']} hrs",        C_AMBER),
    ("Hours Left",     f"{S['hours_remaining']} hrs",       "375623"),
]
col = 1
for label, value, bg in kpis:
    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
    ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
    c4 = ws.cell(row=4, column=col, value=label)
    c5 = ws.cell(row=5, column=col, value=value)
    c4.font = Font("Arial", size=9, bold=True, color=C_WHITE)
    c5.font = Font("Arial", size=18, bold=True, color=C_WHITE)
    for r in (4, 5):
        ws.cell(row=r, column=col).fill      = fill(bg)
        ws.cell(row=r, column=col).alignment = ctr()
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 36
    col += 2

ws.row_dimensions[6].height = 8

# ── Chart data (hidden helper table starting at row 50) ─────────────────────
# Status counts for donut
status_labels = ["Completed", "In Progress", "Blocked", "Pending QA", "Not Started"]
status_values = [S["completed"], S["in_progress"], S["blocked"], S["pending_qa"], S["not_started"]]
for i, (lbl, val) in enumerate(zip(status_labels, status_values)):
    ws.cell(row=50+i, column=20, value=lbl)
    ws.cell(row=50+i, column=21, value=val)

# Burndown data
ws.cell(row=60, column=20, value="Date")
ws.cell(row=60, column=21, value="Remaining")
ws.cell(row=60, column=22, value="Target")
for i, pt in enumerate(burndown):
    ws.cell(row=61+i, column=20, value=pt["date"])
    ws.cell(row=61+i, column=21, value=pt["hours_remaining"])
    ws.cell(row=61+i, column=22, value=pt["target_remaining"])

burn_rows = len(burndown)

# ── Donut chart — task status ─────────────────────────────────────────────────
pie = PieChart()
pie.title    = "Task Status Breakdown"
pie.style    = 10
pie.holeSize = 50
pie.width    = 12
pie.height   = 10
labels = Reference(ws, min_col=20, min_row=50, max_row=54)
data   = Reference(ws, min_col=21, min_row=50, max_row=54)
pie.add_data(data)
pie.set_categories(labels)
pie.series[0].title = None
ws.add_chart(pie, "A7")

# ── Bar chart — estimated vs consumed per task ────────────────────────────────
# We'll populate this from Tasks sheet (referenced cross-sheet after Tasks is built)
# For now create a placeholder — Tasks sheet fills cols 25-27 rows 50+
bar = BarChart()
bar.type    = "bar"
bar.title   = "Hours: Estimated vs. Consumed per Task"
bar.y_axis.title = "Task"
bar.x_axis.title = "Hours"
bar.style   = 10
bar.width   = 20
bar.height  = 14
ws.add_chart(bar, "H7")

# ── Burndown line chart ───────────────────────────────────────────────────────
line = LineChart()
line.title        = "Hours Burndown"
line.y_axis.title = "Hours Remaining"
line.x_axis.title = "Date"
line.style        = 10
line.width        = 14
line.height       = 10

if burn_rows > 0:
    rem_data = Reference(ws, min_col=21, min_row=60, max_row=60+burn_rows)
    tgt_data = Reference(ws, min_col=22, min_row=60, max_row=60+burn_rows)
    line.add_data(rem_data, titles_from_data=True)
    line.add_data(tgt_data, titles_from_data=True)
    line.series[0].title = SeriesLabel(v="Actual Remaining")
    line.series[1].title = SeriesLabel(v="Target")

ws.add_chart(line, "A23")

# Column widths (Dashboard)
for i in range(1, 15):
    ws.column_dimensions[get_column_letter(i)].width = 12
ws.column_dimensions["T"].width = 18
ws.column_dimensions["U"].width = 12
ws.column_dimensions["V"].width = 12

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — TASKS
# ═══════════════════════════════════════════════════════════════════════════════
wt = wb.create_sheet("Tasks")
wt.sheet_properties.tabColor = C_DARK
wt.sheet_view.showGridLines = False

# Header row
wt.merge_cells("A1:L1")
wt["A1"] = "Task Status Board"
wt["A1"].font      = Font("Arial", size=13, bold=True, color=C_WHITE)
wt["A1"].fill      = fill(C_DARK)
wt["A1"].alignment = ctr()
wt.row_dimensions[1].height = 26

cols_t = ["#", "Task Name", "Block", "Est Hrs", "Hrs Used",
          "Overrun %", "Status", "Delay", "Clock Started",
          "Est Completion", "Billing Status", "Notes"]
for ci, h in enumerate(cols_t, 1):
    c = wt.cell(row=2, column=ci, value=h)
    c.font      = Font("Arial", size=9, bold=True, color=C_WHITE)
    c.fill      = fill(C_MID)
    c.alignment = ctr()
    c.border    = TB
wt.row_dimensions[2].height = 24

for ri, t in enumerate(tasks, 3):
    est  = t["est_hours"]
    used = t["hours_used"]
    ovr  = round((used / est - 1) * 100) if est > 0 and used > 0 else 0
    ovr_str = f"{ovr}%" if ovr > 0 else "—"

    # Est completion from clock_started
    est_comp = "—"
    if t["clock_started"]:
        est_comp = f"{t['clock_started']} +{est}h"

    row_data = [
        t["id"], t["name"], t["block"], est, used,
        ovr_str, t["status"], t["delay_status"],
        t["clock_started"] or "—", est_comp,
        t["billing_status"], ""
    ]
    bg = DELAY_COLORS.get(t["delay_status"], "FFFFFF")
    for ci, val in enumerate(row_data, 1):
        c = wt.cell(row=ri, column=ci, value=val)
        c.border    = TB
        c.alignment = ctr() if ci in (1, 3, 4, 5, 6, 7, 8, 9, 10) else lft()
        c.font      = Font("Arial", size=9)
        # Status column coloring
        if ci == 7:
            sc = STATUS_COLORS.get(t["status"], "808080")
            c.fill = fill(sc)
            c.font = Font("Arial", size=9, bold=True, color=C_WHITE)
        elif ci == 11:
            bc = BILLING_COLORS.get(t["billing_status"], "FFFFFF")
            c.fill = fill(bc)
            c.font = Font("Arial", size=9, bold=True)
        elif ci == 6 and ovr > 120:
            c.fill = fill("FCE4D6")
            c.font = Font("Arial", size=9, bold=True, color=C_RED)
        else:
            c.fill = fill(bg)
    wt.row_dimensions[ri].height = 20

col_widths_t = [5, 42, 10, 8, 8, 9, 13, 11, 13, 16, 18, 20]
for i, w in enumerate(col_widths_t, 1):
    wt.column_dimensions[get_column_letter(i)].width = w

wt.freeze_panes = "A3"
wt.auto_filter.ref = f"A2:L{2+len(tasks)}"

# Feed bar chart in Dashboard from this sheet
# Put task ids and hours in hidden cols of Tasks sheet for charting
for ri, t in enumerate(tasks, 3):
    wt.cell(row=ri, column=14, value=f"#{t['id']}")
    wt.cell(row=ri, column=15, value=t["est_hours"])
    wt.cell(row=ri, column=16, value=t["hours_used"])

bar_cats = Reference(wt, min_col=14, min_row=3, max_row=2+len(tasks))
bar_est  = Reference(wt, min_col=15, min_row=3, max_row=2+len(tasks))
bar_used = Reference(wt, min_col=16, min_row=3, max_row=2+len(tasks))
bar.add_data(bar_est)
bar.add_data(bar_used)
bar.set_categories(bar_cats)
bar.series[0].title = SeriesLabel(v="Estimated")
bar.series[1].title = SeriesLabel(v="Consumed")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — BILLING
# ═══════════════════════════════════════════════════════════════════════════════
wb2 = wb.create_sheet("Billing")
wb2.sheet_properties.tabColor = C_GOLD
wb2.sheet_view.showGridLines = False

wb2.merge_cells("A1:J1")
wb2["A1"] = "Billing Tracker"
wb2["A1"].font      = Font("Arial", size=13, bold=True, color=C_WHITE)
wb2["A1"].fill      = fill(C_DARK)
wb2["A1"].alignment = ctr()
wb2.row_dimensions[1].height = 26

cols_b = ["Date", "Task #", "Task Name", "Est Hrs", "Hrs Logged",
          "Overrun Hrs", "Overrun %", "120% Threshold",
          "Approval Status", "Notes"]
for ci, h in enumerate(cols_b, 1):
    c = wb2.cell(row=2, column=ci, value=h)
    c.font      = Font("Arial", size=9, bold=True, color=C_WHITE)
    c.fill      = fill(C_MID)
    c.alignment = ctr()
    c.border    = TB
wb2.row_dimensions[2].height = 24

task_map = {t["id"]: t for t in tasks}
for ri, b in enumerate(billing, 3):
    tid  = b["task_id"]
    tname = task_map[tid]["name"] if tid in task_map else "—"
    ovr_hrs = round(b["hours_logged"] - b["est_hours"], 2)
    threshold = round(b["est_hours"] * 1.2, 2)
    row_data = [
        b["date"], tid, tname, b["est_hours"], b["hours_logged"],
        ovr_hrs, f"{b['overrun_pct']}%", threshold,
        b["approval_status"], ""
    ]
    bg = BILLING_COLORS.get(b["approval_status"], "FFFFFF")
    for ci, val in enumerate(row_data, 1):
        c = wb2.cell(row=ri, column=ci, value=val)
        c.border    = TB
        c.alignment = ctr() if ci in (1, 2, 4, 5, 6, 7, 8) else lft()
        c.font      = Font("Arial", size=9)
        c.fill      = fill(bg)
        if ci == 7 and b["overrun_pct"] > 120:
            c.font = Font("Arial", size=9, bold=True, color=C_RED)
    wb2.row_dimensions[ri].height = 20

col_widths_b = [12, 7, 42, 9, 10, 11, 10, 14, 20, 20]
for i, w in enumerate(col_widths_b, 1):
    wb2.column_dimensions[get_column_letter(i)].width = w

wb2.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — BURNDOWN DATA
# ═══════════════════════════════════════════════════════════════════════════════
wb3 = wb.create_sheet("Burndown Data")
wb3.sheet_properties.tabColor = C_DGREY

wb3["A1"] = "Date"
wb3["B1"] = "Hours Remaining"
wb3["C1"] = "Target Remaining"
for ci in range(1, 4):
    c = wb3.cell(row=1, column=ci)
    c.font = Font("Arial", size=9, bold=True, color=C_WHITE)
    c.fill = fill(C_MID)
    c.alignment = ctr()
    c.border = TB

for ri, pt in enumerate(burndown, 2):
    wb3.cell(row=ri, column=1, value=pt["date"])
    wb3.cell(row=ri, column=2, value=pt["hours_remaining"])
    wb3.cell(row=ri, column=3, value=pt["target_remaining"])
    for ci in range(1, 4):
        wb3.cell(row=ri, column=ci).border = TB
        wb3.cell(row=ri, column=ci).alignment = ctr()

for i, w in enumerate([14, 16, 16], 1):
    wb3.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT_FILE)
print(f"Dashboard written: {OUT_FILE}")
