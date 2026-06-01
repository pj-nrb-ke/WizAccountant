"""Generate QA-Test-003.xlsx from docs/QA/results/qa-cycle-003.json."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
JSON_PATH = ROOT / "docs" / "QA" / "results" / "qa-cycle-003.json"
OUT = ROOT / "QA-Test-003.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
BLOCK_FILL = PatternFill("solid", fgColor="FFEB9C")


def style_header(row):
    for cell in row:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def auto_width(ws):
    for col in ws.columns:
        w = 10
        for cell in col:
            if cell.value is not None:
                w = max(w, min(70, len(str(cell.value)) + 2))
        ws.column_dimensions[col[0].column_letter].width = w


def sheet_data(wb, name, rows, headers):
    ws = wb.create_sheet(name[:31])
    ws.append(headers)
    style_header(ws[1])
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
        status = r.get("Status") or r.get("status")
        row_idx = ws.max_row
        if status == "PASS":
            for c in range(1, len(headers) + 1):
                ws.cell(row_idx, c).fill = PASS_FILL
        elif status == "FAIL":
            for c in range(1, len(headers) + 1):
                ws.cell(row_idx, c).fill = FAIL_FILL
        elif status == "BLOCKED":
            for c in range(1, len(headers) + 1):
                ws.cell(row_idx, c).fill = BLOCK_FILL
    auto_width(ws)


def main():
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    wb = Workbook()
    wb.remove(wb.active)

    sum_ws = wb.create_sheet("Test Summary")
    s = data.get("summary", {})
    for row in [
        ["Cycle", data.get("cycle", "003")],
        ["Date", (data.get("ranAt") or "")[:10]],
        ["Web", data.get("baseUrl")],
        ["API", data.get("apiUrl")],
        ["Instructions", data.get("instructions")],
        ["Overall", s.get("overall")],
        ["Total tests", s.get("total")],
        ["Passed", s.get("passed")],
        ["Failed", s.get("failed")],
        ["Blocked", s.get("blocked")],
        ["Duplicate tests", s.get("duplicateCount")],
        ["Race tests", s.get("raceCount")],
        ["Session tests", s.get("sessionCount")],
        ["Multi-tab tests", s.get("multiTabCount")],
        ["Long-duration tests", s.get("longDurationCount")],
        ["Frontend sync tests", s.get("frontendSyncCount")],
        ["Total nav actions", s.get("totalNavActions")],
    ]:
        sum_ws.append(row)
    style_header(sum_ws[1])
    auto_width(sum_ws)

    cols = ["ID", "Test", "Status", "Assertions", "ms", "Notes", "Evidence", "Reproduction"]
    test_cols = ["id", "test", "status", "assertions", "durationMs", "notes", "evidence", "reproduction"]

    def map_rows(key):
        return [
            {cols[i]: r.get(test_cols[i], "") for i in range(len(cols))}
            for r in data.get(key, [])
        ]

    sheet_data(wb, "Duplicate Prevention Tests", map_rows("duplicatePrevention"), cols)
    sheet_data(wb, "Race Condition Tests", map_rows("raceCondition"), cols)
    sheet_data(wb, "Session Recovery Tests", map_rows("sessionRecovery"), cols)
    sheet_data(wb, "Multi-Tab Tests", map_rows("multiTab"), cols)
    sheet_data(wb, "Long-Duration Stability", map_rows("longDuration"), cols)
    sheet_data(wb, "Frontend Sync Tests", map_rows("frontendSync"), cols)

    ux = wb.create_sheet("UX Findings")
    ux.append(["ID", "Severity", "Finding"])
    style_header(ux[1])
    for u in data.get("uxFindings", []):
        ux.append([u.get("id"), u.get("severity"), u.get("finding")])
    auto_width(ux)

    ev = wb.create_sheet("Evidence Index")
    ev.append(["Artifact", "Path"])
    style_header(ev[1])
    for e in data.get("evidence", []):
        ev.append([e.get("artifact"), e.get("path")])
    auto_width(ev)

    ci = wb.create_sheet("Critical Issues")
    ci.append(["ID", "Severity", "Test", "Notes", "Evidence"])
    style_header(ci[1])
    for c in data.get("criticalIssues", []):
        ci.append([c.get("id"), c.get("severity"), c.get("test"), c.get("notes"), c.get("evidence")])
    auto_width(ci)

    rf = wb.create_sheet("Recommended Fixes")
    rf.append(["ID", "Priority", "Action", "Notes"])
    style_header(rf[1])
    for r in data.get("recommendedFixes", []):
        rf.append([r.get("id"), r.get("priority"), r.get("action"), r.get("notes")])
    auto_width(rf)

    wb.save(OUT)
    print(f"Created {OUT}")


if __name__ == "__main__":
    main()
