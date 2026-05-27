"""Generate WizAccountant.xlsx from Plan 1 phased features."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).resolve().parents[1] / "WizAccountant.xlsx"

# Phase 1 checklist — update `done` as work completes, then run this script.
PHASE1_CHECKLIST = [
    # (id, component, task, done)
    ("P1-01", "Solution", "Create WizAccountant.slnx + Contracts, Api, Connector projects", True),
    ("P1-02", "Contracts", "Shared DTOs: pairing, site, job, heartbeat, run-job message", True),
    ("P1-03", "API", "SQLite persistence (sites, pairing codes, jobs)", True),
    ("P1-04", "API", "POST /api/pairing-codes — generate pairing code", True),
    ("P1-05", "API", "POST /api/sites/pair — register connector site", True),
    ("P1-06", "API", "GET /api/sites — list sites + online status", True),
    ("P1-07", "API", "POST/GET /api/jobs — submit job, poll result", True),
    ("P1-08", "API", "SignalR Connector Hub — register, heartbeat, route RunJob", True),
    ("P1-09", "Connector", "Worker service: pair on first run, save site state", True),
    ("P1-10", "Connector", "Connect to hub, register site, 30s heartbeat loop", True),
    ("P1-11", "Connector", "Receive RunJob, execute handler, POST job result", True),
    ("P1-12", "Connector", "Job shell: Site.Health (mock)", True),
    ("P1-13", "Connector", "Job shell: Customer.List (mock)", True),
    ("P1-14", "Build", "dotnet build WizAccountant.slnx succeeds", True),
    ("P1-15", "Connector", "Wire Pastel.Evolution SDK — Site.Health (live DB)", True),
    ("P1-16", "Connector", "SDK handler: GLAccount.List + GLAccount.Get", True),
    ("P1-17", "Connector", "SDK handler: Customer.List + Customer.Get", True),
    ("P1-18", "Connector", "SDK handler: CustomerTransaction.List (criteria)", True),
    ("P1-19", "Connector", "SDK handler: Supplier.List + SupplierTransaction.List", True),
    ("P1-20", "Connector", "Evolution connection profile + Agent.Authenticate (WizConnector.Setup + DPAPI)", True),
    ("P1-21", "Connector", "WizConnector.Tray — pairing wizard, status, device ID", True),
    ("P1-22", "API", "Auth stub — tenants + users (minimal)", True),
    ("P1-23", "Web", "Admin UI — create pairing code, list sites, Test connection", True),
    ("P1-24", "API", "Job wait/poll helper — sync read with timeout (30–60s)", True),
    ("P1-25", "API", "Basic audit log table for all jobs", True),
    ("P1-26", "Connector", "REST long-poll fallback when WebSocket down", True),
    ("P1-27", "Docs", "Sage connection runbook (DOCS/SAGE-Connection-Process.md); version matrix TBD", True),
    ("P1-28", "Pilot", "End-to-end test on pilot site #1 (pair → online → Customer.List)", True),
    ("P1-29", "Pilot", "End-to-end test on pilot site #2 (second Evolution version)", True),
    ("P1-30", "Success", "Site Online within 60s of connector start", True),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
CHECK_DONE_FILL = PatternFill("solid", fgColor="C6EFCE")
CHECK_TODO_FILL = PatternFill("solid", fgColor="FFC7CE")
CHECK_NEXT_FILL = PatternFill("solid", fgColor="FFEB9C")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PHASE_FILLS = {
    1: PatternFill("solid", fgColor="D6E4F0"),
    2: PatternFill("solid", fgColor="E2EFDA"),
    3: PatternFill("solid", fgColor="FFF2CC"),
    4: PatternFill("solid", fgColor="FCE4D6"),
}


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, max_width=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def add_rows(ws, headers, rows, start_row=1):
    for i, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=i, value=h)
    style_header(ws, start_row)
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
    auto_width(ws)


def main():
    wb = Workbook()

    # --- Overview ---
    ws = wb.active
    ws.title = "Overview"
    overview = [
        ("Product", "WizAccountant — Plan 1"),
        ("Architecture", "On-prem WizConnector (.NET + Sage Evolution SDK) ↔ Cloud WizAccountant (Web + API)"),
        ("Transport", "Outbound HTTPS / WebSocket (no WizVPN, no SSH)"),
        ("System of record", "Client Sage 200 Evolution on LAN (unchanged)"),
        ("Mobile", "iOS + Android — same cloud API as web; no Sage SDK on device"),
        ("Mobile stack (recommended)", "React Native or Flutter; FCM + APNs for push"),
        ("Last updated", "2026-05-26"),
        ("", ""),
        ("Principles", ""),
        ("", "Allowlisted operations only (no raw SDK/SQL from UI or AI)"),
        ("", "Reads before writes; writes need audit + approval from Phase 3"),
        ("", "Every job: jobId, tenantId, siteId; idempotencyKey for posts"),
        ("", "UI shows site online/offline and Sage company name"),
    ]
    ws["A1"] = "Field"
    ws["B1"] = "Value"
    style_header(ws)
    for i, (a, b) in enumerate(overview, 2):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90

    # --- Phase Summary ---
    ws2 = wb.create_sheet("Phase Summary")
    add_rows(
        ws2,
        ["Phase", "Name", "Release", "Duration", "Goal", "Writes", "AI", "Mobile"],
        [
            (1, "Connect & read", "v0.1 Connect", "8–12 weeks", "Pairing, live reads, site online", "No", "No", "No app — API mobile-ready"),
            (2, "Insight", "v0.5 Insight", "8–12 weeks after P1", "Web workspaces + read-only AI", "No", "Read-only tools", "iOS/Android read + push"),
            (3, "Act", "v1.0 Act", "12–16 weeks after P2", "Approvals + posting to live Sage", "Yes (gated)", "Propose → approve", "Approval inbox + push"),
            (4, "Scale", "v1.x Scale", "Ongoing after P3", "Inventory, orders, enterprise", "Expanded", "Workflows", "Practice mode, tablet, store release"),
        ],
    )

    # --- Phase 1 Checklist (early in workbook for visibility) ---
    ws_check = wb.create_sheet("Phase 1 Checklist", 1)
    check_headers = ["Done", "ID", "Component", "Task", "Status"]
    for i, h in enumerate(check_headers, 1):
        ws_check.cell(row=1, column=i, value=h)
    style_header(ws_check)

    next_id = None
    done_count = 0
    for row_idx, (task_id, component, task, done) in enumerate(PHASE1_CHECKLIST, 2):
        mark = "☑" if done else "☐"
        status = "Complete" if done else "Pending"
        if not done and next_id is None:
            next_id = task_id
            status = "NEXT"

        ws_check.cell(row=row_idx, column=1, value=mark)
        ws_check.cell(row=row_idx, column=2, value=task_id)
        ws_check.cell(row=row_idx, column=3, value=component)
        ws_check.cell(row=row_idx, column=4, value=task)
        ws_check.cell(row=row_idx, column=5, value=status)

        row_fill = CHECK_DONE_FILL if done else (CHECK_NEXT_FILL if status == "NEXT" else CHECK_TODO_FILL)
        for col in range(1, 6):
            ws_check.cell(row=row_idx, column=col).fill = row_fill
        if done:
            done_count += 1

    total = len(PHASE1_CHECKLIST)
    ws_check.cell(row=total + 3, column=1, value="Summary")
    ws_check.cell(row=total + 3, column=4, value=f"{done_count} / {total} complete")
    if next_id:
        next_task = next(t for t in PHASE1_CHECKLIST if t[0] == next_id)
        ws_check.cell(row=total + 4, column=1, value="Next task")
        ws_check.cell(row=total + 4, column=2, value=next_id)
        ws_check.cell(row=total + 4, column=4, value=next_task[2])

    ws_check.freeze_panes = "A2"
    ws_check.column_dimensions["A"].width = 8
    ws_check.column_dimensions["B"].width = 10
    ws_check.column_dimensions["C"].width = 14
    ws_check.column_dimensions["D"].width = 72
    ws_check.column_dimensions["E"].width = 12

    next_task_id = next_id
    next_task_text = next(t[2] for t in PHASE1_CHECKLIST if t[0] == next_id) if next_id else ""

    # --- All Features (master sheet) ---
    features = []

    def feat(phase, component, area, feature, details, rw="Read"):
        features.append((phase, component, area, feature, details, rw))

    # Phase 1 - Connector
    for f, d in [
        ("Windows Service", "Auto-start, recovery, structured logging"),
        ("System tray app", "Pairing wizard, connection status, device/site ID, open logs"),
        ("Pairing", "One-time code from cloud → store siteId + credentials (DPAPI)"),
        ("Cloud link", "Outbound TLS; WebSocket/SignalR + REST fallback long-poll"),
        ("Heartbeat", "Every 30–60s; version, SDK/Evolution version, company DB name"),
        ("Local config", "Evolution connection profile, SDK agent user (encrypted)"),
        ("Health handler", "Site.Health — DB reachable, SDK auth OK"),
    ]:
        feat(1, "WizConnector", "Platform", f, d)

    for domain, ops in [
        ("GL", "GLAccount.List, GLAccount.Get"),
        ("AR", "Customer.List, Customer.Get, CustomerTransaction.List"),
        ("AP", "Supplier.List, Supplier.Get, SupplierTransaction.List"),
        ("Reports", "Trial balance–style extracts (scoped queries)"),
        ("Inventory (optional)", "InventoryItem.List, InventoryItem.Get"),
    ]:
        feat(1, "WizConnector", "Sage SDK", domain, ops)

    for f, d in [
        ("Auth", "Email/password or SSO stub; tenants + users"),
        ("Sites", "Register tenant → pairing code → list sites, last seen"),
        ("Connector Hub", "Route jobs to siteId; job queue + timeout"),
        ("Admin UI", "Sites list, connection status, Test connection"),
        ("Job API", "Submit job → wait/poll result (sync read 30–60s cap)"),
        ("Audit (basic)", "Log job type, user, site, timestamp, success/fail"),
    ]:
        feat(1, "WizAccountant Cloud", "Platform", f, d)

    # Phase 2
    for f, d in [
        ("Expanded reads", "Outstanding balances, aged analysis, detail by Autoidx"),
        ("Search / filter", "Criteria builders: account, date range, reference"),
        ("Performance", "Connection pooling, timeouts, pagination"),
        ("Diagnostics", "Support bundle: redacted logs, versions, last 50 jobs"),
        ("Auto-update (optional)", "Signed installer check channel"),
    ]:
        feat(2, "WizConnector", "Platform", f, d)

    for domain, ops in [
        ("AR", "Open items/outstanding; allocations read"),
        ("AP", "Same for suppliers"),
        ("GL", "Transaction list by account/period"),
        ("Inventory", "Stock levels, item search"),
        ("Orders", "SalesOrder.List, PurchaseOrder.List"),
        ("Master data", "Projects, warehouses, tax rates, transaction codes list"),
    ]:
        feat(2, "WizConnector", "Sage SDK", domain, ops)

    for f, d in [
        ("Dashboard", "TB summary, debtors/creditors KPIs"),
        ("AR/AP workspaces", "Browse customers/suppliers, drill to transactions"),
        ("Global search", "Cross-entity search"),
        ("AI chat (read-only)", "Tools = read handlers; citations"),
        ("Conversation history", "Per tenant/user"),
        ("Export", "CSV/PDF for visible lists"),
        ("Notifications", "Email: invites, site offline (e.g. Brevo)"),
    ]:
        feat(2, "WizAccountant Cloud", "Platform", f, d)

    for f, d in [
        ("Tool allowlist", "Read handlers only"),
        ("System prompt", "Never claim posted; suggest approval for actions"),
        ("Data timestamp", "Show data as of job completion"),
    ]:
        feat(2, "WizAccountant Cloud", "AI Guardrails", f, d)

    for f, d in [
        ("Auth", "Login, biometric unlock, secure token storage"),
        ("Sites", "List sites, online/offline, switch tenant/site"),
        ("Dashboard", "Debtors/creditors totals, TB snapshot (read-only)"),
        ("AR/AP quick view", "Customer/supplier search, open items"),
        ("Push notifications", "Site offline alert; optional daily summary"),
        ("AI chat (read-only)", "Same tools as web; mobile UI"),
        ("Deep links", "Open entity from notification"),
    ]:
        feat(2, "WizAccountant Mobile", "iOS / Android", f, d)

    # Phase 3
    for f, d in [
        ("Idempotency store", "Local SQLite: idempotencyKey → result hash"),
        ("Write handlers", "Allowlisted post operations"),
        ("Transactions", "BeginTran / Commit / Rollback per action"),
        ("StartNewBatch", "Separate audit batches per post"),
        ("Error mapping", "Evolution exceptions → stable error codes"),
        ("Consent (optional)", "Tray: Allow cloud to post for 1 hour"),
    ]:
        feat(3, "WizConnector", "Platform", f, d, "Write")

    for domain, ops, notes in [
        ("GL", "GLTransaction.Post; journal batch create only", "VAT legs per SDK"),
        ("AR", "CustomerTransaction.Post; Allocations.Save", "Codes per site config"),
        ("AP", "SupplierTransaction.Post; Allocations.Save", "Codes per site config"),
        ("Master (limited)", "Customer.Save, Supplier.Save", "Role-gated"),
        ("Orders (optional 3b)", "SalesOrder/PurchaseOrder Save + Process partial", "Sub-phase"),
    ]:
        feat(3, "WizConnector", "Sage SDK", domain, f"{ops} — {notes}", "Write")

    for f, d in [
        ("Approval queue", "Proposed journal/payment → approver → execute"),
        ("Maker-checker", "Prepare vs approve roles"),
        ("AI propose, human approve", "Draft only until approved"),
        ("Write audit", "Before/after, user, approver, Evolution ref"),
        ("Workflows (starter)", "Month-end checklist, payment run proposal"),
        ("Site config sync", "Transaction codes, tax, branch"),
        ("Rollback messaging", "Posts not auto-reversed in product"),
    ]:
        feat(3, "WizAccountant Cloud", "Platform", f, d, "Write")

    for f, d in [
        ("Approval inbox", "Pending journals, payments, allocations"),
        ("Approve / reject", "Maker-checker with comment"),
        ("Push notifications", "Approval required + deep link"),
        ("AI drafts", "View proposal; approve via flow"),
        ("Audit (read)", "Approval history per item"),
    ]:
        feat(3, "WizAccountant Mobile", "iOS / Android", f, d, "Write")

    # Phase 4
    for f, d in [
        ("Multi-company", "One agent → multiple company DBs"),
        ("Branch context", "SetBranchContext per site"),
        ("Heavy modules", "CN, RTS, inventory adj, warehouse transfer"),
        ("Order processing", "SO/PO Process, partial, Complete"),
        ("Offline queue (optional)", "Buffer approved jobs when Sage busy"),
    ]:
        feat(4, "WizConnector", "Platform", f, d, "Write")

    for f, d in [
        ("Practice mode", "One firm, many client sites"),
        ("RBAC", "Read vs propose vs approve vs admin"),
        ("SSO", "Azure AD / Google"),
        ("Monitoring", "Site SLA, failed job alerts"),
        ("Billing", "Per site / per user hooks"),
        ("Compliance", "DPA templates, retention"),
        ("Advanced AI", "Multi-step workflows; doc upload draft"),
    ]:
        feat(4, "WizAccountant Cloud", "Platform", f, d)

    for domain, ops in [
        ("Inventory", "InventoryTransaction.Post; CN/RTS Process; warehouse"),
        ("Orders", "Full SO/PO lifecycle"),
        ("CRM", "Incidents (if needed)"),
        ("Job costing", "Job card Save only (no process)"),
    ]:
        feat(4, "WizConnector", "Sage SDK", domain, ops, "Write")

    for f, d in [
        ("Practice mode", "Switch client site for accounting firms"),
        ("Inventory (read)", "Stock lookup, low-stock push alert"),
        ("Workflows", "Starter checklist steps from phone"),
        ("Tablet layout", "iPad / Android tablet for approvers"),
        ("App store release", "Production signing, crash analytics, staged rollout"),
    ]:
        feat(4, "WizAccountant Mobile", "iOS / Android", f, d)

    feat(1, "WizAccountant Cloud", "Platform", "API mobile-ready", "JWT, same job endpoints — no native app in P1")

    ws3 = wb.create_sheet("Features")
    add_rows(
        ws3,
        ["Phase", "Component", "Area", "Feature", "Details", "Read/Write"],
        features,
    )
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        phase = row[0].value
        if phase in PHASE_FILLS:
            for cell in row:
                cell.fill = PHASE_FILLS[phase]

    # --- Success Criteria ---
    ws4 = wb.create_sheet("Success Criteria")
    criteria = [
        (1, "Pair reads on 2 pilot sites (Evolution versions documented)"),
        (1, "Site Online within 60s of service start"),
        (1, "Read customer list + AR open items without desktop Sage"),
        (1, "No inbound ports on customer firewall"),
        (2, "Natural language → correct AR/AP answer on pilot data"),
        (2, "5 read tools in chat with >99% success rate"),
        (2, "Accountants prefer browser for enquiry vs Evolution"),
        (2, "Mobile: login, dashboard, one AR enquiry on iOS or Android"),
        (3, "GL journal: propose → approve → balanced post in live Sage"),
        (3, "Duplicate idempotency key does not double-post"),
        (3, "Allocation payment to invoice on pilot AR"),
        (3, "Full audit trail export for one transaction"),
        (3, "Mobile: approver completes one payment from push notification"),
        (4, "20+ sites under management with monitoring"),
        (4, "At least one inventory workflow in production"),
        (4, "Enterprise pilot with SSO + practice multi-site"),
    ]
    add_rows(ws4, ["Phase", "Success criterion"], criteria)

    # --- Cross-phase Platform ---
    ws5 = wb.create_sheet("Platform (Cross-Phase)")
    add_rows(
        ws5,
        ["Capability", "Phase introduced"],
        [
            ("Shared contracts NuGet (JobRequest / JobResult)", 1),
            ("PostgreSQL (tenants, sites, jobs, audit)", 1),
            ("SignalR / WebSocket hub", 1),
            ("Serilog + correlation id", 1),
            ("DPAPI secrets on connector", 1),
            ("CI: connector MSI + API docker", 1),
            ("Evolution version matrix doc", 1),
            ("Rate limiting per site", 2),
            ("API versioning", 2),
            ("Push notifications (FCM / APNs)", 2),
            ("Mobile apps (iOS + Android)", 2),
            ("Mobile approval UX", 3),
        ],
    )

    # --- Mobile Strategy ---
    ws_mobile = wb.create_sheet("Mobile Strategy")
    mobile_rows = [
        ("Principle", "Mobile uses same WizAccountant cloud API as web — no WizConnector or Sage SDK on device"),
        ("Phase 1", "No native app; ensure API auth (JWT), pagination, mobile-friendly JSON"),
        ("Phase 2", "Ship iOS + Android: read-only parity with key web screens + push"),
        ("Phase 3", "Approval inbox — primary mobile value for managers on the go"),
        ("Phase 4", "Practice multi-site, tablet UX, app store hardening"),
        ("Recommended stack", "React Native or Flutter (one codebase); alternatively .NET MAUI if team is C#-only"),
        ("Push", "Firebase Cloud Messaging (Android) + Apple Push Notification service"),
        ("Security", "Biometric unlock, short-lived tokens, certificate pinning (enterprise option)"),
        ("Out of scope", "Offline posting to Sage; full Evolution parity on mobile"),
    ]
    ws_mobile["A1"] = "Topic"
    ws_mobile["B1"] = "Detail"
    style_header(ws_mobile)
    for i, (a, b) in enumerate(mobile_rows, 2):
        ws_mobile.cell(row=i, column=1, value=a)
        ws_mobile.cell(row=i, column=2, value=b)
    ws_mobile.column_dimensions["A"].width = 22
    ws_mobile.column_dimensions["B"].width = 85

    # --- Module Priority ---
    ws6 = wb.create_sheet("Sage Module Priority")
    add_rows(
        ws6,
        ["Priority", "Module", "Phases", "Notes"],
        [
            (1, "GL", "1–3", "Accounts, journals"),
            (2, "AR / AP", "1–3", "Masters, transactions, allocations"),
            (3, "Inventory", "2 read, 4 write", "Items, adjustments"),
            (4, "Sales / Purchase orders", "3b / 4", "High complexity"),
            (5, "CRM / Job costing", "4+", "Niche; job card no process via SDK"),
        ],
    )

    # --- Out of scope ---
    ws7 = wb.create_sheet("Not Building (Plan 1)")
    items = [
        "WizVPN / network tunnel to LAN",
        "Hosted customer DB restore as primary product",
        "Full cloud replacement of Evolution UI (Plan 3)",
        "Raw SQL access from web or AI",
        "Evolution report printing via SDK",
        "Unrestricted SDK passthrough",
        "Cashbook / journal batch process via SDK",
        "Job card process/complete via SDK",
    ]
    ws7["A1"] = "Item"
    style_header(ws7)
    for i, item in enumerate(items, 2):
        ws7.cell(row=i, column=1, value=item)
    auto_width(ws7)

    # --- Phase 1 Out of scope ---
    ws8 = wb.create_sheet("Phase 1 Exclusions")
    add_rows(
        ws8,
        ["Excluded from Phase 1"],
        [(x,) for x in [
            "AI chat",
            "Workflows",
            "Writes to ledger",
            "Multi-site analytics",
            "Native mobile app (API only in P1)",
            "Inventory documents",
            "Orders Process()",
            "Allocations",
        ]],
    )

    # --- SDK deferred (Phase 3) ---
    ws9 = wb.create_sheet("SDK Limits (Phase 3+)")
    add_rows(
        ws9,
        ["Limitation", "Notes"],
        [
            ("Cashbook / journal batch process", "SDK cannot process — create only"),
            ("Job card process/complete", "Not supported by SDK"),
            ("Evolution printing", "Not available in SDK"),
        ],
    )

    # --- Execution Status (synced from checklist) ---
    ws10 = wb.create_sheet("Execution Status")
    exec_rows = []
    for task_id, component, task, done in PHASE1_CHECKLIST:
        exec_rows.append(
            (
                "Done" if done else ("Next" if task_id == next_task_id else "Pending"),
                f"{task_id}: {task}",
                component,
                "2026-05-26",
            )
        )
    add_rows(ws10, ["Status", "Work item", "Component", "Updated (UTC)"], exec_rows)
    ws10.column_dimensions["A"].width = 12
    ws10.column_dimensions["B"].width = 80
    ws10.column_dimensions["C"].width = 14
    ws10.column_dimensions["D"].width = 16

    wb.save(OUT)
    print(f"Created {OUT}")


if __name__ == "__main__":
    main()
